
import openpyxl
import PySimpleGUI as sg
from datetime import datetime


# =========================================================

# given key already exists in a dictionary.
def check_key(dict: object, key: object) -> object:
    if key in dict.keys():
        return True
    else:
        return False


# =========================================================
class MngEmployee:

    def __init__(self, num='', name='', goals='', report_to=''):
        self.num = num
        self.name = name
        self.goals = goals
        self.report_to = report_to
        self.done = 0
        self.total = 0
        self.num_done = 0
        self.num_not_done = 0
        self.my_employees = {}

    def set_is_done(self):
        if self.goals == "Edit this goal to document your 2019 Development Plan.":
            self.done = 0
        else:
            self.done = 1

    def add_employee(self, employee):
        if not check_key(self.my_employees, employee.name):
            self.my_employees[employee.name] = employee

    def __repr__(self):
        return "Name: " + str(self.name) + " * Total: " + str(self.total) + " * Done: " + \
               str(self.num_done) + " * Not Done: " + str(self.num_not_done) + "\n"

    def sum(self):
        for key, value in self.my_employees.items():
            e = self.my_employees[key]
            self.total = self.total + 1
            if e.done == 1:
                self.num_done = self.num_done + 1
            else:
                self.num_not_done = self.num_not_done + 1


# =========================================================


class Employee:
    def __init__(self, num='', name='', goals='', report_to=''):
        self.num = num
        self.name = name
        self.goals = goals
        self.report_to = report_to
        self.done = 0

    def set_is_done(self):
        if self.goals == "Edit this goal to document your 2019 Development Plan.":
            self.done = 0
        else:
            self.done = 1

    def __repr__(self):
        return "{Employee: " + str(self.num) + " * " + str(self.name) + " * " + str(self.report_to) + " * " + \
               str(self.goals) + " * " + str(self.done) + "}" + "\n"


# =========================================================

def analyse(all_dict: object, mng_dict: object) -> object:
    for key, value in all_dict.items():
        e = all_dict[key]
        e.set_is_done()

        if not check_key(mng_dict, e.report_to):
            mng_dict[e.report_to] = MngEmployee()

        mngEmployee = mng_dict[e.report_to]
        mngEmployee.name = e.report_to
        mngEmployee.add_employee(e)


# =========================================================


def read(file_name, my_dict):
    # open a excel file with .xlsx format
    excelfile = openpyxl.load_workbook(file_name)

    # get names of all spreadsheet in the file
    # print(excelfile.get_sheet_names())

    # get the first spreadsheet by name
    # sheet1 = excelfile.get_sheet_by_name("Page1")
    sheet1 = excelfile.get_active_sheet()

    # get the number of rows in the sheet
    # print(sheet1.max_row)

    # get the number of columns in the sheet
    # print(sheet1.max_column)

    # read the whole spreadsheet
    for row_index in range(2, (sheet1.max_row + 1)):

        employee_num = sheet1.cell(row=row_index, column=1).value
        if not check_key(my_dict, employee_num):
            my_dict[employee_num] = Employee(employee_num)

        e = my_dict[employee_num]
        e.name = sheet1.cell(row=row_index, column=6).value
        e.report_to = sheet1.cell(row=row_index, column=8).value
        e.goals = sheet1.cell(row=row_index, column=17).value


# =========================================================
if __name__ == "__main__":
    # empty dictionary
    all_dict = {}
    mng_dict = {}

    event, (file_name,) = sg.Window('Hadas is the best').Layout([[sg.Text('Document to open')],
                                                     [sg.In(), sg.FileBrowse()],
                                                     [sg.OK('Open'), sg.CloseButton('Cancel')]]).Read()
    # layout = [[sg.Text('Input XSLX file'), [sg.In(), sg.FileBrowse()]],
    #          [sg.OK()]]

    # window = sg.Window('Hadas is the best').Layout(layout)
    # event, values = window.Read()
    # file_name = values[0]

    # read the xlsx file

    read(file_name, all_dict)

    # set if done or not based on the goal string
    analyse(all_dict, mng_dict)

    # print("=======================================")
    # sagi = mng_dict["Dudai, Sagi"]
    # print(sagi)

    # print("=======================================")

    wb = openpyxl.load_workbook(file_name)
    num_of_sheets = len(wb.get_sheet_names())
    sum_sheet_name = "sum_" + str(datetime.now().time().microsecond)
    wb.create_sheet(sum_sheet_name,num_of_sheets)
    sheet = wb.get_sheet_by_name(sum_sheet_name)

    index = 1
    sheet['A' + str(index)].value = "Name"
    sheet['C' + str(index)].value = "Total"
    sheet['D' + str(index)].value = "Done"
    sheet['E' + str(index)].value = "Not Done"

    index = index + 1
    for key, value in all_dict.items():

        if value.report_to == "Dudai, Sagi" and  value.name in mng_dict:
            mng_dict[value.name].sum()
            # print(mng_dict[value.name])

            sheet['A' + str(index)].value = mng_dict[value.name].name
            sheet['C' + str(index)].value = mng_dict[value.name].total
            sheet['D' + str(index)].value = mng_dict[value.name].num_done
            sheet['E' + str(index)].value = mng_dict[value.name].num_not_done

            index = index + 1
            for key_e, value_e in mng_dict[value.name].my_employees.items():
                if value_e.name in mng_dict:
                    mng_dict[value_e.name].sum()
                    # print(mng_dict[value_e.name])

                    sheet['B' + str(index)].value = mng_dict[value_e.name].name
                    sheet['C' + str(index)].value = mng_dict[value_e.name].total
                    sheet['D' + str(index)].value = mng_dict[value_e.name].num_done
                    sheet['E' + str(index)].value = mng_dict[value_e.name].num_not_done

                    index = index + 1

    wb.save(file_name)
    layout_done = [[sg.Text('Done')],
                   [sg.OK()]]

    window_done = sg.Window('Hadas is the best').Layout(layout_done)
    event_done, values_done = window_done.Read()