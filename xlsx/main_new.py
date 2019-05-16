
import openpyxl
import PySimpleGUI as sg
from datetime import datetime
# import networkx as nx
# =========================================================

# given key already exists in a dictionary.
def check_key(dict: object, key: object) -> object:
    if key in dict.keys():
        return True
    else:
        return False


# =========================================================
class MngEmployee:

    def __init__(self, num='', name='', goals='', report_to=''):
        self.num = num
        self.name = name
        self.goals = goals
        self.report_to = report_to
        self.done = 0
        self.total = 0
        self.num_done = 0
        self.num_not_done = 0
        self.my_employees = {}
        self.my_managers = {}

    def set_is_done(self):
        if self.goals == "Edit this goal to document your 2019 Development Plan.":
            self.done = 0
        else:
            self.done = 1

    def add_manager(self, manager):
        if not check_key(self.my_managers, manager.name):
            self.my_managers[manager.name] = manager

    def add_employee(self, employee):
        if not check_key(self.my_employees, employee.name):
            self.my_employees[employee.name] = employee

    def __repr__(self):
        return "Name: " + str(self.name) + " * Total: " + str(self.total) + " * Done: " + \
               str(self.num_done) + " * Not Done: " + str(self.num_not_done) + "\n"

    def sum(self):
        for key, value in self.my_employees.items():
            e = self.my_employees[key]
            self.total = self.total + 1
            if e.done == 1:
                self.num_done = self.num_done + 1
            else:
                self.num_not_done = self.num_not_done + 1


# =========================================================


class Employee:
    def __init__(self, num='', name='', goals='', report_to=''):
        self.num = num
        self.name = name
        self.goals = goals
        self.report_to = report_to
        self.done = 0

    def set_is_done(self):
        if self.goals == "Edit this goal to document your 2019 Development Plan.":
            self.done = 0
        else:
            self.done = 1

    def __repr__(self):
        return "{Employee: " + str(self.num) + " * " + str(self.name) + " * " + str(self.report_to) + " * " + \
               str(self.goals) + " * " + str(self.done) + "}" + "\n"

# =========================================================


def build_org_chart(all_d: object, m_dict: object, o_chart: object):

    # find root - report col is empty
    for k, v in all_d.items():
        e = all_d[k]
        o_chart.add_node(e.name)

    for k, v in all_d.items():
        e = all_d[k]
        o_chart.add_edge(e.report_to, e.name)

    print(list(nx.connected_components(o_chart)))

# =========================================================


def analyse(all_d: object, m_dict: object) -> object:
    for k, v in all_d.items():
        e = all_d[k]
        e.set_is_done()

        if not check_key(m_dict, e.report_to):
            m_dict[e.report_to] = MngEmployee()
            m_dict[e.report_to].name = e.report_to

        m_dict[e.report_to].add_employee(e)

    for mk, mv in m_dict.items():
        if check_key(all_d, mk) and all_d[mk].report_to != '':
           mv.report_to = all_d[mk].report_to


# =========================================================


def read(f_name: object, my_dict: object) -> object:
    # open a excel file with .xlsx format
    excel_file = openpyxl.load_workbook(f_name)

    # get names of all spreadsheet in the file
    # print(excel_file.get_sheet_names())

    # get the first spreadsheet by name
    # sheet1 = excel_file.get_sheet_by_name("Page1")
    sheet1 = excel_file.get_active_sheet()

    # get the number of rows in the sheet
    # print(sheet1.max_row)

    # get the number of columns in the sheet
    # print(sheet1.max_column)

    # read the whole spreadsheet
    for row_index in range(2, (sheet1.max_row)):

        employee_num = sheet1.cell(row=row_index, column=1).value
        employee_name = sheet1.cell(row=row_index, column=6).value
        if not check_key(my_dict, employee_name):
            my_dict[employee_name] = Employee(employee_num, employee_name)
            my_dict[employee_name].report_to = sheet1.cell(row=row_index, column=8).value
            my_dict[employee_name].goals = sheet1.cell(row=row_index, column=17).value


# =========================================================
def calc(f_name, m_dict, all_d ):
    wb = openpyxl.load_workbook(f_name)
    num_of_sheets = len(wb.get_sheet_names())
    sum_sheet_name = "sum_" + str(datetime.now().time().microsecond)
    wb.create_sheet(sum_sheet_name, num_of_sheets)
    sheet = wb.get_sheet_by_name(sum_sheet_name)

    index = 1
    sheet['A' + str(index)].value = "Name"
    sheet['C' + str(index)].value = "Total"
    sheet['D' + str(index)].value = "Done"
    sheet['E' + str(index)].value = "Not Done"

    index = index + 1
    for key, value in all_d.items():

        # if value.report_to == "Dudai, Sagi" and value.name in mng_dict:
        if value.name in m_dict:
            m_dict[value.name].sum()
            # print(mng_dict[value.name])

            sheet['A' + str(index)].value = m_dict[value.name].name
            sheet['C' + str(index)].value = m_dict[value.name].total
            sheet['D' + str(index)].value = m_dict[value.name].num_done
            sheet['E' + str(index)].value = m_dict[value.name].num_not_done

            index = index + 1
            for key_e, value_e in m_dict[value.name].my_employees.items():
                if value_e.name in m_dict:
                    m_dict[value_e.name].sum()
                    # print(mng_dict[value_e.name])

                    sheet['B' + str(index)].value = m_dict[value_e.name].name
                    sheet['C' + str(index)].value = m_dict[value_e.name].total
                    sheet['D' + str(index)].value = m_dict[value_e.name].num_done
                    sheet['E' + str(index)].value = m_dict[value_e.name].num_not_done

                    index = index + 1

    wb.save(f_name)


# =========================================================


g_row_index = 2

def the_recursive(sheet, mng_dic, curr_dic, col_index, alignment, total_mng):

    global g_row_index
    for key, value in curr_dic.items():

        if check_key(mng_dic, value.name):

            mng_dic[value.name].sum()

            sheet.cell(row=g_row_index, column=col_index + 1).value = mng_dic[value.name].name

            sheet.cell(row=g_row_index, column=col_index + 2 + alignment).value = mng_dic[value.name].total

            sheet.cell(row=g_row_index, column=col_index + 3 + alignment).value = mng_dic[value.name].num_done

            sheet.cell(row=g_row_index, column=col_index + 4 + alignment).value = mng_dic[value.name].num_not_done

            sheet.cell(row=g_row_index, column=col_index + 5 + alignment).value = (int(mng_dic[value.name].num_done) / int(mng_dic[value.name].total)) * 100

            g_row_index = g_row_index + 1

            # create new entry in the manager array
            if not check_key(total_mng, value.name):
                total_mng[value.name] = MngEmployee()
                total_mng[value.name].name = value.name

            if not check_key(total_mng, mng_dic[value.name].report_to):
                total_mng[value.name].add_manager(mng_dic[value.name])

            the_recursive(sheet, mng_dic, mng_dic[value.name].my_employees, col_index+1, alignment-1, total_mng)


# =========================================================


def calc_new(f_name, m_dict, total_mng):
    wb = openpyxl.load_workbook(f_name)
    num_of_sheets = len(wb.get_sheet_names())
    sum_sheet_name = "sum_" + str(datetime.now().time().microsecond)
    wb.create_sheet(sum_sheet_name, num_of_sheets)
    sheet = wb.get_sheet_by_name(sum_sheet_name)

    '''root = None
    for key, value in m_dict.items():
        if value.report_to == '':
            root = value
    '''
    root = m_dict["Dudai, Sagi"]
    if root:
        the_recursive(sheet, m_dict, root.my_employees, 0, 3, total_mng)
    wb.save(f_name)


# =========================================================
if __name__ == "__main__":
    # empty dictionary
    all_dict = {}
    mng_dict = {}
    total_mng = {}

    # org_chart = nx.Graph()

    event, (file_name,) = sg.Window('Hadas is the best').Layout([[sg.Text('Document to open')],
                                                     [sg.In(), sg.FileBrowse()],
                                                     [sg.OK('Open'), sg.CloseButton('Cancel')]]).Read()
    # layout = [[sg.Text('Input XSLX file'), [sg.In(), sg.FileBrowse()]],
    #          [sg.OK()]]

    # window = sg.Window('Hadas is the best').Layout(layout)
    # event, values = window.Read()
    # file_name = values[0]

    # read the xlsx file
    if file_name is None or not file_name != "":
        exit(0)
    read(file_name, all_dict)

    # set if done or not based on the goal string
    analyse(all_dict, mng_dict)

    #build org chart
    # build_org_chart(all_dict, mng_dict, org_chart)

    # print("=======================================")
    # sagi = mng_dict["Dudai, Sagi"]
    # print(sagi)

    # print("=======================================")

    calc_new(file_name,mng_dict, total_mng)

    print(total_mng)
    #calc(file_name, all_dict, mng_dict)

    layout_done = [[sg.Text('Done')],
                   [sg.OK()]]

    window_done = sg.Window('Hadas is the best').Layout(layout_done)
    event_done, values_done = window_done.Read()