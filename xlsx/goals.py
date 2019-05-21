
import openpyxl
import PySimpleGUI as sg
from datetime import datetime
# =========================================================

g_row_index = 2


# =========================================================
# given key already exists in a dictionary.
def check_key(dict: object, key: object) -> object:
    if key in dict.keys():
        return True
    else:
        return False


# =========================================================


class Employee:
    def __init__(self, num='', name='', goals='', report_to=''):
        self.num = num
        self.name = name
        self.goals = goals
        self.report_to = report_to
        self.done = 0
        self.total = 0
        self.num_done = 0
        self.num_not_done = 0
        self.my_employees = {}
        self.is_manager = 0
        self.my_employees_total = 0
        self.my_employees_done = 0
        self.my_employees_not_done = 0

    # =========================================================
    def set_is_done(self):
        if self.goals == "Edit this goal to document your 2019 Development Plan.":
            self.done = 0
        else:
            self.done = 1

    # =========================================================
    def add_employee(self, employee):
        if not check_key(self.my_employees, employee.name):
            self.my_employees[employee.name] = employee

    # =========================================================
    def sum(self):
        for key, value in self.my_employees.items():
            e = self.my_employees[key]
            self.total = self.total + 1
            if e.done == 1:
                self.num_done = self.num_done + 1
            else:
                self.num_not_done = self.num_not_done + 1

    # =========================================================
    def my_employees_sum(self):

        for key, value in self.my_employees.items():
            if value.is_manager:
                self.my_employees_total = self.my_employees_total + value.total
                if value.done == 1:
                    self.my_employees_done = self.my_employees_done + value.num_done
                else:
                    self.my_employees_not_done = self.my_employees_not_done + value.num_not_done
        
        # At least one of the employee is also a manager
        if self.my_employees_total != 0:
            self.my_employees_total = self.my_employees_total + self.total
            self.my_employees_done = self.my_employees_done + self.num_done
            self.my_employees_not_done = self.my_employees_not_done + self.num_not_done

    # =========================================================
    def __repr__(self):
        return "{Employee: " + str(self.num) + " * " + str(self.name) + " * " + str(self.report_to) + " * " + \
               str(self.goals) + " * " + str(self.done) + "}" + "\n"

    # =========================================================
    def save_to_file(self,sheet, col_index, alignment):
        global g_row_index

        # Adding blank row between each manager only the first level
        if col_index == 1:
            g_row_index = g_row_index + 1
        # Save my total info
        sheet.cell(row=g_row_index, column=col_index + 1).value = self.name
        if int(self.my_employees_total):
            sheet.cell(row=g_row_index, column=col_index + 2 + alignment).value = self.my_employees_total

            sheet.cell(row=g_row_index, column=col_index + 3 + alignment).value = self.my_employees_done

            sheet.cell(row=g_row_index, column=col_index + 4 + alignment).value = self.my_employees_not_done

            sheet.cell(row=g_row_index, column=col_index + 5 + alignment).value = round((int(
                self.my_employees_done) / int(self.my_employees_total)) * 100)

            g_row_index = g_row_index + 1

        # Save my info
        sheet.cell(row=g_row_index, column=col_index + 2 + alignment).value = self.total

        sheet.cell(row=g_row_index, column=col_index + 3 + alignment).value = self.num_done

        sheet.cell(row=g_row_index, column=col_index + 4 + alignment).value = self.num_not_done

        sheet.cell(row=g_row_index, column=col_index + 5 + alignment).value = round(
            (int(self.num_done) / int(self.total)) * 100)

        # my employees info
        for e_name, employee in self.my_employees.items():
            if employee.is_manager:
                g_row_index = g_row_index + 1
                employee.save_to_file(sheet,col_index+1, alignment-1)


# =========================================================


def analyse(all_d: object) -> object:
    for k, v in all_d.items():
        e = all_d[k]
        e.set_is_done()

        if e and e.report_to and check_key(all_d,e.report_to):
            all_d[e.report_to].add_employee(e)
            all_d[e.report_to].is_manager = 1

# =========================================================


def calc(all_d: object) -> object:
    for k, v in all_d.items():
        e = all_d[k]
        if e.is_manager:
            e.sum()

    for ke, ve in all_d.items():
        ee = all_d[ke]
        if ee.is_manager:
            ee.my_employees_sum()

# =========================================================


def read(f_name: object, my_dict: object) -> object:
    # open a excel file with .xlsx format
    excel_file = openpyxl.load_workbook(f_name)

    # get names of all spreadsheet in the file
    # print(excel_file.get_sheet_names())

    sheet1= excel_file.worksheets[0]
    # get the first spreadsheet by name
    # sheet1 = excel_file.get_sheet_by_name("Page1")
    # sheet1 = excel_file.get_active_sheet()

    # get the number of rows in the sheet
    # print(sheet1.max_row)

    # get the number of columns in the sheet
    # print(sheet1.max_column)

    # read the whole spreadsheet
    for row_index in range(2, sheet1.max_row):

        employee_num = sheet1.cell(row=row_index, column=1).value
        employee_name = sheet1.cell(row=row_index, column=6).value
        if not check_key(my_dict, employee_name):
            my_dict[employee_name] = Employee(employee_num, employee_name)
            my_dict[employee_name].report_to = sheet1.cell(row=row_index, column=8).value
            my_dict[employee_name].goals = sheet1.cell(row=row_index, column=17).value


# =========================================================
def write_to_file(f_name, a_dict, r_name):
    wb = openpyxl.load_workbook(f_name)
    num_of_sheets = len(wb.get_sheet_names())
    sum_sheet_name = "sum_" + str(datetime.now().time().microsecond)
    wb.create_sheet(sum_sheet_name, num_of_sheets)
    sheet = wb.get_sheet_by_name(sum_sheet_name)

    root = a_dict[r_name]
    if root:
        root.save_to_file(sheet, 0, 3)
    wb.save(f_name)


# =========================================================
if __name__ == "__main__":
    # empty dictionary
    all_dict = {}
    root_name = "Dudai, Sagi"

    event, (file_name,) = sg.Window('Hadas is the best').Layout([[sg.Text('Document to open')],
                                                     [sg.In(), sg.FileBrowse()],
                                                     [sg.OK('Open'), sg.CloseButton('Cancel')]]).Read()

    # read the xlsx file
    if file_name is None or not file_name != "":
        exit(0)
    read(file_name, all_dict)

    # set if done or not based on the goal string
    analyse(all_dict)

    calc(all_dict)

    write_to_file(file_name, all_dict, root_name)

    layout_done = [[sg.Text('Done')],
                   [sg.OK()]]

    window_done = sg.Window('Hadas is the best').Layout(layout_done)
    event_done, values_done = window_done.Read()